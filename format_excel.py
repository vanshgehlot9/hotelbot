import openpyxl

# Load original
wb = openpyxl.load_workbook('Jodhpur_Hotels.xlsx')
ws = wb.active

# Create new workbook
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# Required headers
headers = ["name", "city", "price_per_night", "rating", "amenities", "description", "tenant_id"]
new_ws.append(headers)

# Read from original and write to new
for row in ws.iter_rows(min_row=2, values_only=True):
    hotel_name = row[0]
    district = row[1]
    mobile = row[2]
    address = row[3]
    
    if not hotel_name:
        continue

    # Map to new format with defaults
    name = str(hotel_name)
    city = str(district) if district else "Jodhpur"
    price = 3000  # default
    rating = 4.5  # default
    amenities = "WiFi, AC, TV"
    description = f"Located at {address}. Contact: {mobile}"
    tenant_id = "platform" # Default to superadmin tenant for now

    new_ws.append([name, city, price, rating, amenities, description, tenant_id])

new_wb.save("Formatted_Jodhpur_Hotels.xlsx")
print("Successfully created Formatted_Jodhpur_Hotels.xlsx")
