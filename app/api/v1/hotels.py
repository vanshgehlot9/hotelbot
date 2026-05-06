from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from app.api.dependencies import get_firestore_db, require_superadmin, require_tenant_admin
import openpyxl
import io
import uuid

router = APIRouter()

@router.get("/")
async def list_hotels(
    db=Depends(get_firestore_db),
    current_user: dict = Depends(require_superadmin)
):
    """List all hotels across all tenants."""
    docs = db.collection("hotels").stream()
    hotels = []
    for d in docs:
        hotel_data = d.to_dict()
        hotel_data["id"] = d.id
        hotels.append(hotel_data)
    return {"hotels": hotels}

@router.delete("/{hotel_id}")
async def delete_hotel(
    hotel_id: str,
    db=Depends(get_firestore_db),
    current_user: dict = Depends(require_superadmin)
):
    """Delete a hotel by ID."""
    doc_ref = db.collection("hotels").document(hotel_id)
    doc = doc_ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Hotel not found")
    
    doc_ref.delete()
    return {"status": "ok", "msg": f"Hotel {hotel_id} deleted successfully"}

@router.post("/upload-excel")
async def upload_hotels_excel(
    file: UploadFile = File(...),
    db=Depends(get_firestore_db),
    current_user: dict = Depends(require_superadmin)
):
    """
    Upload an Excel file to bulk-import hotels into Firestore.
    Expected columns (row 1 = header):
    name | city | price_per_night | rating | amenities | description | tenant_id
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    required = {"name", "city", "price_per_night", "rating", "amenities", "description", "tenant_id"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")

    def idx(col): return headers.index(col)

    from app.core.security import get_password_hash
    
    added = 0
    errors = []
    generated_accounts = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            name = str(row[idx("name")]).strip()
            city = str(row[idx("city")]).strip()
            
            # Auto-generate credentials for the hotel
            slug = name.lower().replace(' ', '').replace(',', '').replace('-', '').replace('.', '')
            generated_email = f"admin@{slug}.com"
            generated_password = "password123"
            new_tenant_id = f"tenant_{uuid.uuid4().hex[:8]}"
            
            if not name or not city:
                continue

            # 1. Create Tenant Admin User
            user_ref = db.collection("users").where("email", "==", generated_email).get()
            if not user_ref: # Only create if email doesn't exist
                db.collection("users").add({
                    "email": generated_email,
                    "hashed_password": get_password_hash(generated_password),
                    "role": "tenant_admin",
                    "tenant_id": new_tenant_id,
                    "is_active": True
                })
                generated_accounts.append(f"{name}: {generated_email} / {generated_password}")
            else:
                # If user already exists, grab their existing tenant_id
                new_tenant_id = user_ref[0].to_dict().get("tenant_id", new_tenant_id)

            # 2. Create Hotel
            doc_id = f"{name.lower().replace(' ', '_').replace(',', '')}_{uuid.uuid4().hex[:4]}"
            data = {
                "name": name,
                "city": city,
                "city_lower": city.lower(),
                "price_per_night": int(float(row[idx("price_per_night")] or 0)),
                "rating": float(row[idx("rating")] or 0),
                "amenities": str(row[idx("amenities")] or ""),
                "description": str(row[idx("description")] or ""),
                "tenant_id": new_tenant_id,
                "available": True,
            }
            if "state" in headers:
                data["state"] = str(row[idx("state")]).strip()
                
            db.collection("hotels").document(doc_id).set(data)
            added += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    # We can add generated accounts to the errors/warnings just so the frontend displays them in the yellow box!
    if generated_accounts:
        errors.append("--- GENERATED ACCOUNTS ---")
        errors.extend(generated_accounts)

    return {
        "status": "ok",
        "hotels_added": added,
        "errors": errors
    }
